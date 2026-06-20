"""
Generates gstr2b_may2026.xlsx — mimics the actual GST portal GSTR-2B export format.
Run from project root:  python testing/generate_gstr2b_xlsx.py
Requires: pip install openpyxl
"""

from pathlib import Path
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUT_PATH = Path(__file__).parent / "gstr2b_may2026.xlsx"

# ── colour palette matching GST portal export ──────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "DEEAF1"
LIGHT_GREY = "F2F2F2"
WARN_ORANGE= "FFE0B2"
RED_BG     = "FFCCCC"
GREEN_BG   = "E2EFDA"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(color="000000", size=10):
    return Font(bold=True, color=color, size=size)

def header_font():
    return Font(bold=True, color="FFFFFF", size=9)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def wrap_left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — GSTR-2B Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "GSTR-2B Summary"

# Portal header bar
ws1.merge_cells("A1:H1")
ws1["A1"] = "GSTR-2B — Auto-Drafted Input Tax Credit Statement"
ws1["A1"].font = Font(bold=True, color="FFFFFF", size=13)
ws1["A1"].fill = fill(DARK_BLUE)
ws1["A1"].alignment = wrap_center()
ws1.row_dimensions[1].height = 28

meta = [
    ("GSTIN",            "27ACCPS7562H1Z8"),
    ("Legal Name",       "Suryakant Optics"),
    ("Trade Name",       "Suryakant Optics"),
    ("Return Period",    "May 2026  (05/2026)"),
    ("Generation Date",  "14-06-2026"),
    ("Status",           "Generated"),
]
for i, (k, v) in enumerate(meta, start=2):
    ws1[f"A{i}"] = k
    ws1[f"B{i}"] = v
    ws1[f"A{i}"].font = bold_font(DARK_BLUE)
    ws1[f"B{i}"].font = Font(size=10)
    ws1[f"A{i}"].fill = fill(LIGHT_BLUE)
    ws1[f"B{i}"].fill = fill(LIGHT_GREY)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 38

# ITC Summary table
row = len(meta) + 3
ws1[f"A{row}"] = "ITC AVAILABILITY SUMMARY — May 2026"
ws1[f"A{row}"].font = bold_font("FFFFFF", 11)
ws1[f"A{row}"].fill = fill(MID_BLUE)
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"].alignment = wrap_center()
row += 1

itc_headers = ["Supplier", "GSTIN", "Invoice Total (₹)", "ITC Available (₹)"]
for col, h in enumerate(itc_headers, 1):
    c = ws1.cell(row, col, h)
    c.font = header_font()
    c.fill = fill(MID_BLUE)
    c.alignment = wrap_center()
    c.border = thin_border()
row += 1

itc_rows = [
    ("Nashik Optical Wholesalers Pvt Ltd",   "27AABCN5432P1Z3", 25200.00,   2700.00,  GREEN_BG),
    ("SafeVision Lens Manufacturing Co",     "33AADCS7891K1Z2", 21280.00,   2280.00,  GREEN_BG),
    ("Royal Optica Frames",                  "24AABCR8765H1Z4", "NOT FILED","NOT FILED",RED_BG),
    ("Woodland Residency",                   "27AADCW3456P1Z8", 4130.00,       0.00,  WARN_ORANGE),
    ("Shree Vinayak Spectacle Works",        "27ABCPV6789H1Z1", 28320.00,   2880.00,  WARN_ORANGE),
    ("VisionTech Optical Solutions",         "27AADCV3421P1ZN", 476000.00, 51000.00,  RED_BG),
    ("Titan Eyeplus Distribution Network",   "29AABCT8765P1Z4", 40320.00,   4320.00,  GREEN_BG),
    ("Arihant Optics Ltd",                   "27AABCA9876H1Z7", 42560.00,   4560.00,  LIGHT_BLUE),
]

for sup, gstin, total, itc, bg in itc_rows:
    for col, val in enumerate([sup, gstin, total, itc], 1):
        c = ws1.cell(row, col, val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = wrap_left() if col <= 2 else wrap_center()
        if col in (3, 4) and isinstance(val, float):
            c.number_format = '₹#,##0.00'
    row += 1

# Totals row
ws1.cell(row, 1, "TOTAL").font = bold_font()
ws1.cell(row, 1).fill = fill(LIGHT_BLUE)
ws1.cell(row, 3, 637610.00).number_format = '₹#,##0.00'
ws1.cell(row, 3).font = bold_font()
ws1.cell(row, 3).fill = fill(LIGHT_BLUE)
ws1.cell(row, 4, "₹67,740  (excl. Royal Optica & restaurant)").font = bold_font()
ws1.cell(row, 4).fill = fill(LIGHT_BLUE)
for col in range(1, 5):
    ws1.cell(row, col).border = thin_border()

ws1.column_dimensions["A"].width = 40
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 22

# Notes
row += 2
notes = [
    ("⚠️  IMPORTANT NOTES", ""),
    ("Royal Optica Frames has NOT filed GSTR-1 for May 2026.", "ITC of ₹3,840 is AT RISK — contact supplier urgently"),
    ("Woodland Residency (restaurant bill) — ITC BLOCKED under Sec 17(5)(b).", "Do not claim ₹630"),
    ("Shree Vinayak Spectacle Works — HSN rate mismatch (18% charged, 12% correct).", "Request credit note for ₹1,440 excess tax"),
    ("VisionTech Optical Solutions — GSTIN registered Feb 2026 (<180 days).", "High-value invoice ₹4.76L — verify before claiming ITC"),
    ("Arihant Optics Ltd — filed in GSTR-2B but no invoice scanned.", "MISSED ITC of ₹4,560 — locate physical invoice"),
]
for note, fix in notes:
    ws1[f"A{row}"] = note
    ws1[f"C{row}"] = fix
    ws1[f"A{row}"].font = Font(italic=True, size=9, color="C00000" if "⚠️" not in note else "7F3F00")
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — B2B Inward Supplies (raw data, exactly like GST portal)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("B2B Inward Supplies")

# Title
ws2.merge_cells("A1:O1")
ws2["A1"] = "B2B Inward Supplies — GSTR-2B May 2026"
ws2["A1"].font = Font(bold=True, color="FFFFFF", size=11)
ws2["A1"].fill = fill(DARK_BLUE)
ws2["A1"].alignment = wrap_center()
ws2.row_dimensions[1].height = 22

b2b_cols = [
    "GSTIN of Supplier", "Trade/Legal Name", "Invoice Number",
    "Invoice Date", "Invoice Value (₹)", "Place of Supply",
    "Reverse Charge", "Invoice Type",
    "Rate (%)", "Taxable Value (₹)",
    "IGST (₹)", "CGST (₹)", "SGST (₹)", "CESS (₹)",
    "ITC Available",
]
for col, h in enumerate(b2b_cols, 1):
    c = ws2.cell(2, col, h)
    c.font = header_font()
    c.fill = fill(MID_BLUE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws2.row_dimensions[2].height = 40

b2b_data = [
    ("27AABCN5432P1Z3","Nashik Optical Wholesalers Pvt Ltd","NOL/2025-26/1847","08-05-2026",25200.00,"Maharashtra (27)","N","Regular",12,22500.00,0.00,1350.00,1350.00,0.00,"YES",GREEN_BG),
    ("33AADCS7891K1Z2","SafeVision Lens Manufacturing Co",  "SVLC/MAY/2026/892","14-05-2026",21280.00,"Maharashtra (27)","N","Regular",12,19000.00,2280.00,0.00,0.00,0.00,"YES",GREEN_BG),
    ("27AADCW3456P1Z8","Woodland Residency",                "WR/MAY/2026/4521","22-05-2026", 4130.00,"Maharashtra (27)","N","Regular",18, 3500.00,0.00,315.00,315.00,0.00,"YES*",WARN_ORANGE),
    ("27ABCPV6789H1Z1","Shree Vinayak Spectacle Works",     "SVSW/2526/344",   "05-05-2026",28320.00,"Maharashtra (27)","N","Regular",18,24000.00,0.00,2160.00,2160.00,0.00,"YES**",WARN_ORANGE),
    ("27AADCV3421P1ZN","VisionTech Optical Solutions",      "VTOS/2026/001",   "25-05-2026",476000.00,"Maharashtra (27)","N","Regular",12,425000.00,0.00,25500.00,25500.00,0.00,"YES***",RED_BG),
    ("29AABCT8765P1Z4","Titan Eyeplus Distribution Network","TITANB00040320",  "12-05-2026",40320.00,"Maharashtra (27)","N","Regular",12,36000.00,4320.00,0.00,0.00,0.00,"YES",GREEN_BG),
    ("27AABCA9876H1Z7","Arihant Optics Ltd",                "AOL/2526/0847",   "02-05-2026",42560.00,"Maharashtra (27)","N","Regular",12,38000.00,0.00,2280.00,2280.00,0.00,"YES",LIGHT_BLUE),
]

money_cols = {5, 10, 11, 12, 13, 14}  # 1-indexed

for data_row in b2b_data:
    row_idx = ws2.max_row + 1
    *vals, bg = data_row
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row_idx, col, val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = wrap_center() if col not in (1, 2, 3) else wrap_left()
        if col in money_cols and isinstance(val, float):
            c.number_format = '#,##0.00'

# Footnote
fn_row = ws2.max_row + 2
footnotes = [
    "*  Woodland Residency: ITC legally BLOCKED under Section 17(5)(b) — food & beverage services",
    "** Shree Vinayak: HSN rate mismatch — invoice shows 18%, correct rate is 12% for HSN 9004. Request credit note.",
    "*** VisionTech: GSTIN registered 14-Feb-2026 (<180 days). High-value invoice — verify supplier authenticity.",
    "NOTE: Royal Optica Frames (GSTIN 24AABCR8765H1Z4) has NOT filed GSTR-1 for May 2026. ITC of ₹3,840 is AT RISK.",
]
for fn in footnotes:
    ws2.cell(fn_row, 1, fn).font = Font(italic=True, size=8, color="7F3F00")
    ws2.merge_cells(f"A{fn_row}:O{fn_row}")
    fn_row += 1

# column widths
col_widths = [22, 38, 20, 14, 14, 20, 12, 12, 8, 16, 12, 12, 12, 10, 14]
for col, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Munim Analysis (bonus sheet for CA)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Munim ITC Analysis")

ws3.merge_cells("A1:F1")
ws3["A1"] = "Munim.ai — ITC Analysis & Action Items  |  May 2026  |  Suryakant Optics"
ws3["A1"].font = Font(bold=True, color="FFFFFF", size=11)
ws3["A1"].fill = fill(DARK_BLUE)
ws3["A1"].alignment = wrap_center()
ws3.row_dimensions[1].height = 24

analysis_headers = ["Invoice", "Supplier", "Amount (₹)", "ITC Status", "ITC (₹)", "Action Required"]
for col, h in enumerate(analysis_headers, 1):
    c = ws3.cell(2, col, h)
    c.font = header_font()
    c.fill = fill(MID_BLUE)
    c.alignment = wrap_center()
    c.border = thin_border()

analysis_data = [
    ("NOL/2025-26/1847", "Nashik Optical Wholesalers",    25200.00,  "✅ CONFIRMED",        2700.00,  "None — ITC ready to claim",                         GREEN_BG),
    ("SVLC/MAY/2026/892","SafeVision Lens Mfg Co",        21280.00,  "⚠️  PROBABLE MATCH",  2280.00,  "Invoice no. format differs (SVLC/.../00892 vs /892). Verify with supplier.",LIGHT_BLUE),
    ("ROF/2526/MAY/0156","Royal Optica Frames",            35840.00,  "🔴 AT RISK",          3840.00,  "URGENT: Supplier has NOT filed GSTR-1 for May 2026. Call supplier immediately.", RED_BG),
    ("WR/MAY/2026/4521", "Woodland Residency",              4130.00,  "🚫 INELIGIBLE",          0.00,  "Restaurant bill — Section 17(5)(b). No ITC available. Cost to P&L.",         WARN_ORANGE),
    ("SVSW/2526/344",    "Shree Vinayak Spectacle Works",  28320.00,  "🔧 FIXABLE",          2880.00,  "HSN 9004 @ 18% (wrong). Correct is 12%. Get credit note for ₹1,440 excess.", WARN_ORANGE),
    ("VTOS/2026/001",    "VisionTech Optical Solutions",  476000.00,  "🚨 FRAUD FLAG",          0.00,  "HOLD: GSTIN <180 days old, invoice #001, value ₹4.76L. Verify before claim.", RED_BG),
    ("TET/NAS/0512/2026","Titan Eyeplus Distribution",     40320.00,  "🟡 POSSIBLE MATCH",   4320.00,  "Invoice no. format differs from GSTR-2B (TITANB00040320). Verify once.",       LIGHT_BLUE),
    ("AOL/2526/0847",    "Arihant Optics Ltd",             42560.00,  "💸 MISSED ITC",       4560.00,  "GSTR-2B has ₹4,560 ITC but invoice never scanned! Locate and upload.",        LIGHT_BLUE),
]

for row_data in analysis_data:
    row_idx = ws3.max_row + 1
    *vals, bg = row_data
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row_idx, col, val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = wrap_left()
        if col in (3, 5) and isinstance(val, float):
            c.number_format = '#,##0.00'

# Totals
total_row = ws3.max_row + 1
ws3.cell(total_row, 1, "TOTAL / OPPORTUNITY").font = bold_font()
ws3.cell(total_row, 1).fill = fill(LIGHT_BLUE)
ws3.cell(total_row, 3, 673610.00).number_format = '#,##0.00'
ws3.cell(total_row, 3).font = bold_font()
ws3.cell(total_row, 3).fill = fill(LIGHT_BLUE)
ws3.cell(total_row, 5, 25300.00).number_format = '#,##0.00'
ws3.cell(total_row, 5).font = bold_font()
ws3.cell(total_row, 5).fill = fill(LIGHT_BLUE)
ws3.cell(total_row, 6, "Confirmed ₹9,300 + Recoverable ₹5,280 + Missed ₹4,560 = ₹19,140 total ITC opportunity").font = bold_font()
ws3.cell(total_row, 6).fill = fill(LIGHT_BLUE)
for col in range(1, 7):
    ws3.cell(total_row, col).border = thin_border()

col_widths3 = [22, 38, 16, 20, 14, 65]
for col, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

wb.save(OUT_PATH)
print(f"✅ XLSX generated: {OUT_PATH}")
print(f"   Sheets: {[s.title for s in wb.worksheets]}")
print(f"   Size  : {OUT_PATH.stat().st_size:,} bytes")
